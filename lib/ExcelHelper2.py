
import openpyxl
from mytools import basicClass
from mytools import init_chdir


# 版本说明：2021.02.20 新建
# 功能说明：
# 备注说明：示例https://www.cnblogs.com/gdjlc/p/11407587.html
# 环境要求：pip install openpyxl
# 设备要求：


class Excel_helper(basicClass):
    """docstring for Excel_reader
    这个是一个excel助手，有读取方法，采用的是openpyxl模块"""
    def __init__(self, name='excel助手openpyxl:'):
        super(Excel_helper, self).__init__()
        self.instanceName = name
        self.openFileName = ''

    def openXlsx(self, dir_file):
        self.workbook = openpyxl.load_workbook(dir_file)  # 创建一个excel对象
        self.sheets = []
        for sheet in self.workbook:
            self.sheets.append(sheet)  # 返回一个工作表对象列表（所有工作表）
        # self.sheets = self.workbook.sheets()
        self.openFileName = dir_file.split("/")[-1]  # 获取文件名
        self.dir_file = dir_file
        self.sheetNames = self.workbook.sheetnames

    def read(self, sheet_index, rowx, colx):
        '''
        :param sheet_index:int
        :param rowx:int
        :param colx:int
        读取某个工作表中的一个单元格的值，传入的参数为工作表序号，行，列
        '''
        if rowx >= 0 and colx >= 0:
            celldata = self.sheets[sheet_index].cell(rowx+1, colx+1).value
            self.classPrint("%s文件,%s表,%s行,%s列,数据%s" %
                            (self.openFileName, sheet_index, rowx, colx, celldata))
            # table.cell(rowx, colx)  # 返回单元格对象
            # table.cell_type(rowx, colx)  # 返回单元格中的数据类型
            # table.cell_value(rowx,colx)   #返回单元格中的数据
            return celldata
        else:
            return None

    def write(self, sheet_index, rowx, colx, setValue):
        '''
        :param sheet_index:int
        :param rowx:int
        :param colx:int
        :param setValue:str
        写入某个工作表中的一个单元格的值，传入的参数为工作表序号，行，列，值
        '''
        self.classPrint("在{}中{}工作表的{}行{}列写入{}".format(
                                                    self.openFileName,
                                                    self.sheetNames[sheet_index],
                                                    rowx,
                                                    colx,
                                                    setValue
                                                     ))
        self.sheets[sheet_index].cell(rowx+1, colx+1, value=setValue)
        try:
            self.workbook.save(self.dir_file)
        except Exception as e:
            self.classPrint(f'保存失败{self.dir_file}', e)
            return False
        else:
            return True

    def getSheetNrows(self, sheet_index=0, echo=0):  # 读取某个工作表中的有内容的总行数
        if echo != 0:
            print(f"{self.openFileName} 的{sheet_index}表总共有{self.sheets[sheet_index].max_row}行")
            # print("%s总共有%s行" % (self.openFileName, self.sheets[sheet_index].max_row))
        return self.sheets[sheet_index].max_row

    def switchSheet(self, sheetId):
        '''
        切换到某个工作表
        '''
        # table = data.sheet_by_index(0)  # 通过索引获取，例如打开第一个sheet表格
        # table = data.sheet_by_name("sheet1")  # 通过名称获取，如读取sheet1表单
        if sheetId in range(0, len(self.sheets)):
            self.sheet = self.sheets[sheetId]  # 通过索引顺序获取
            return self.sheet
        return None
        # 以上三个函数都会返回一个xlrd.sheet.Sheet()对象

    def getAllSheetNames(self, echo=0):
        '''
        # 返回book中所有工作表的名字
        '''
        self.sheetNames = self.workbook.sheetnames
        if echo != 0:
            self.classPrint("共有%s个工作表,名字分别为%s" % (len(self.sheetNames), self.sheetNames))
        return self.sheetNames

    # 2.2 对列进行操作：
    def getSheetCol(self, sheetId, y):
        '''
        获取某表的某列,返回xlrd.sheet.Cell的list
        param sheetId:int
        param y:int
        '''
        self.classPrint("获取%s文件,%s表%s列" % (self.openFileName, sheetId, y))
        i = list(self.sheets[sheetId].columns)[y]
        return i

    # 2.2 对行进行操作：
    def getSheetRow(self, sheetId, x, echo=0):
        '''
        获取某表的某行,返回xlrd.sheet.Cell的list
        param sheetId:int
        param x:int
        '''
        if echo != 0:
            self.classPrint("获取%s文件,%s表%s行" % (self.openFileName, sheetId, x))
        i = list(self.sheets[sheetId].rows)[x]  # TODO
        return i

    def getColId(self, sheetId, rowx, byName):
        """获取byName在rowx中对应的列号

        Args:
            sheetId (int): 工作表下标数
            rowx (int): 要在第rowx行查找
            byName (str): 要查找的字符串

        Returns:
            [int]: 返回查找到的列数
        """
        row0 = self.getSheetRow(sheetId, rowx)
        for colId, cell in enumerate(row0):
            if isinstance(cell.value, type("str")) and byName in cell.value:  # 1 为string也就是text
                self.classPrint("找到包含'%s'的列号%s,值为:%s" % (byName, colId, cell.value))
                return colId
        return False  # 到这一步就是没有找到，有找到就早就找到了

    def getCellValueType(self, sheetId, rowx, colx):
        # 对单元格进行操作：
        cell = self.sheets[sheetId].cell(rowx+1, colx+1).value
        # ctype = typelist[str(cell)]
        print("type==", type(cell))
        return type(cell)

    def getSubject2Dict(self, sheetId=0, colId=0):  # 获取第一个工作表中的第一行转换成字典, {subject: colno}
        dict1 = {}
        for name, colno in enumerate(self.getSheetCol(sheetId=sheetId, y=colId)):
            print(name, colno)
            dict1[name] = colno
        return dict1
        pass

    """ 读取sheet对象中的日期 """
    #  TODO
    def getCellDate(self, sheetId, row, col):
        pass

    def getColDuplication(self, sheetId, y):
        '''
        判断某表某列中是否有value重复,
        '''
        col = self.getSheetCol(sheetId, y)
        for num in range(0, len(col)):
            for n in range(num+1, len(col)):
                if col[num].value == col[n].value:
                    result = "发现%s行中数据项重复%s,请手动去除item%s" % (num, n, col[num].value)
                    self.oneAlarmInfo = result
                    return False
        return True

    def saveXlsx(self, xlsxfileName):
        print('111111111111111111')
        self.classPrint("将%s保存为%s" % (self.openFileName, xlsxfileName))
        try:
            self.workbook.save(xlsxfileName)
        except Exception:
            self.classPrint('保存{}失败'.format(xlsxfileName))
            return False
        return True

    def setCheckDict(self, dict):
        self.dict = dict

    def checkCol1(self,):
        for dictKey in self.dict:
            if self.getColId(0, 0, dictKey) == self.dict[dictKey]:
                self.classPrint(dictKey, '检查通过')


if __name__ == '__main__':
    print("I'm test.py")
    # 初始化一下工作目录
    init_chdir()
    xlsxFile_ZXG = 'demo-ZXG.xlsx'
    xlsxFile_supplier = 'demo-supplier.xlsx'

    # Gracy是苏菲亚公主里面的冰雪女巫
    Gracy = Excel_helper()
    Gracy.setNickname('Gracy:')
    Gracy.openXlsx(xlsxFile_ZXG)
    data11 = Gracy.read(0, 1, 1)
    print("data11", type(data11))
    Gracy.getAllSheetNames()
    Gracy.read(0, 0, 0)
    Gracy.getCellValueType(0, 1, 1)
    print(Gracy.getColDuplication(0, 3))
    print(Gracy.oneAlarmInfo)
    p = Gracy.switchSheet(1)
    i = p.max_column
    print("i", i)
    # 采用这样的方法进行数据的写入
    Gracy.write(1, 0, 0, "hahahah")
    # 写入完成后另存为一个xlsx文件
    Gracy.saveXlsx('saveDemo.xlsx')
